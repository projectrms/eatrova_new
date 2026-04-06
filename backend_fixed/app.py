# app.py
from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_socketio import SocketIO, emit, join_room, leave_room
# import sqlite3
from datetime import datetime
from flask import send_file
from openpyxl import Workbook
import os
import csv, io
from flask import Response
from werkzeug.utils import secure_filename
import stripe
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from flask import send_file
import io

stripe.api_key = os.getenv("STRIPE_API_KEY")  # load from .env

UPLOAD_FOLDER = "static/menu"
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}

# -------------------------
# Config
# -------------------------
# DB_PATH = os.path.join(os.getcwd(), "restaurant.db")

app = Flask(__name__)

CORS(
    app,
    resources={r"/*": {"origins": "http://localhost:5173"}},
    supports_credentials=True
)
# Use eventlet for websocket support (ensure eventlet is installed)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="eventlet")
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
activity_logs = []

TAX_CONFIG = {
    "gst": 0.05,       # 5%
    "service": 0.05    # 2%
}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# from dotenv import load_dotenv
# load_dotenv()
# -------------------------
# DB helpers
# -------------------------
import psycopg2
from psycopg2.extras import RealDictCursor

DATABASE_URL = os.getenv("DATABASE_URL")

def get_db_connection():
    return psycopg2.connect(
        DATABASE_URL,
        cursor_factory=RealDictCursor
    )

def init_db():
    conn = get_db_connection()
    cur = conn.cursor()

    # Users (customers)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            name TEXT,
            email TEXT UNIQUE
        )
    """)

    # Chefs
    cur.execute("""
        CREATE TABLE IF NOT EXISTS chefs (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    """)

    # Waiters
    cur.execute("""
        CREATE TABLE IF NOT EXISTS waiters (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    """)

    # Managers
    cur.execute("""
        CREATE TABLE IF NOT EXISTS managers (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    """)

    # Menu items
    cur.execute("""
        CREATE TABLE IF NOT EXISTS menu_items (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT,
            price NUMERIC(10,2) NOT NULL,
            image TEXT,
            category TEXT DEFAULT 'main',
            available BOOLEAN DEFAULT TRUE
        )
    """)

    # Tables
    cur.execute("""
        CREATE TABLE IF NOT EXISTS tables (
            id SERIAL PRIMARY KEY,
            table_number INTEGER UNIQUE NOT NULL,
            active BOOLEAN DEFAULT TRUE
        )
    """)

    # Orders
    cur.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id),
            table_id INTEGER NOT NULL REFERENCES tables(id),
            table_no INTEGER NOT NULL,
            customer_name TEXT DEFAULT '',
            total NUMERIC(10,2) NOT NULL,
            paid BOOLEAN DEFAULT FALSE,
            status TEXT DEFAULT 'pending',
            created_at TIMESTAMP NOT NULL,
            updated_at TIMESTAMP,
            cancel_reason TEXT
        )
    """)

    # Order items
    cur.execute("""
        CREATE TABLE IF NOT EXISTS order_items (
            id SERIAL PRIMARY KEY,
            order_id INTEGER NOT NULL REFERENCES orders(id),
            menu_id INTEGER NOT NULL REFERENCES menu_items(id),
            quantity INTEGER NOT NULL,
            price NUMERIC(10,2) NOT NULL
        )
    """)

    conn.commit()
    cur.close()
    conn.close()

def seed_menu_items():
    conn = get_db_connection()
    cur = conn.cursor()
    count = cur.execute("SELECT COUNT(*) FROM menu_items").fetchone()[0]
    if count == 0:
        items = [
            ("Margherita Pizza", "Classic cheese and tomato pizza", 299, "https://i.imgur.com/MK7wGkV.png"),
            ("Veg Burger", "Loaded with veggies and cheese", 199, "https://i.imgur.com/xKz3XjZ.png"),
            ("Pasta Alfredo", "Creamy white sauce pasta", 249, "https://i.imgur.com/XU2vUeG.png"),
            ("Cold Coffee", "Iced coffee with cream", 149, "https://i.imgur.com/EVp5pGf.png"),
        ]
        cur.executemany("INSERT INTO menu_items (name, description, price, image) VALUES (%s, %s, %s, %s)", items)
        conn.commit()
    conn.close()


# ---------------------------
# Helper functions for sockets
# ---------------------------
def seed_tables():
    conn = get_db_connection()
    cur = conn.cursor()

    count = cur.execute("SELECT COUNT(*) FROM tables").fetchone()[0]

    if count == 0:
        for i in range(1, 11):  # tables 1 to 10
            cur.execute(
                "INSERT INTO tables (table_number, active) VALUES (%s, 1)",
                (i,)
            )

    conn.commit()
    conn.close()

def get_unpaid_count(user_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM orders WHERE user_id = %s AND paid = 0", (user_id,))
        row = cur.fetchone()
        conn.close()
        return int(row[0]) if row else 0
    except Exception as e:
        print("get_unpaid_count error:", e)
        return 0


def emit_unpaid_update(user_id):
    try:
        count = get_unpaid_count(user_id)
        payload = {"user_id": user_id, "unpaid_count": count}

        # compatibility events
        socketio.emit("unpaid_bill_update", payload)
        socketio.emit("unpaid_updated", payload)

        # user-specific rooms
        socketio.emit("unpaid_bill_update", payload, room=f"user_{user_id}")
        socketio.emit("unpaid_updated", payload, room=f"user_{user_id}")
    except Exception as e:
        print("emit_unpaid_update error:", e)


def send_bill_update(user_id):
    """Send bill_update event to user's room."""
    socketio.emit("bill_update", {"user_id": user_id}, room=f"user_{user_id}")

#Log Function 
def log_event(message):
    entry = {
        "message": message,
        "time": datetime.now().strftime("%I:%M %p")
    }
    activity_logs.append(entry)
    socketio.emit("activity_event", entry, broadcast=True)

# ---------------------------
# Socket event handlers
# ---------------------------
@socketio.on("order_created_client")
def handle_client_order_created(data):
    user_id = data.get("user_id")
    if user_id:
        emit_unpaid_update(user_id)


@socketio.on("connect")
def on_connect():
    print("Socket connected:", request.sid)
    emit("connected", {"msg": "connected"})


@socketio.on("disconnect")
def on_disconnect():
    print("Socket disconnected:", request.sid)

@socketio.on("leave_room")
def handle_leave_room(data):
    room = data.get("room")
    if room:
        leave_room(room)
        emit("left_room", {"room": room})
        print(f"SID {request.sid} left room {room}")


@socketio.on("bill_paid")
def handle_bill_paid(data):
    user_id = data.get("user_id")
    socketio.emit("bill_update", room=f"user_{user_id}")


@socketio.on("subscribe_user")
def handle_subscribe_user(data):
    user_id = data.get("user_id")
    if user_id:
        room = f"user_{user_id}"
        join_room(room)
        emit("subscribed", {"room": room})
        print(f"SID {request.sid} subscribed to {room}")

@socketio.on("join_customer")
def handle_join_customer(data):
    user_id = data if isinstance(data, int) else data.get("user_id") or data.get("id")
    if user_id:
        room = f"user_{user_id}"
        join_room(room)
        emit("joined_customer", {"room": room})
        print(f"SID {request.sid} joined customer room {room}")

@socketio.on("join")
def handle_join(data):
    room = data.get("room")
    join_room(room)
    print(f"Client joined room: {room}")

@socketio.on("waiter_update")
def waiter_update(data):
    socketio.emit("order_updated", data, room="chef")

#Log Fetch Api
@app.route("/logs", methods=["GET"])
def get_logs():
    return jsonify(activity_logs)

# ---------------------------
# HTTP ROUTES
# ---------------------------
@app.route("/")
def home():
    return "✅ Backend running (Flask + SQLite + SocketIO)."


from flask import send_from_directory

@app.route("/static/menu/<path:filename>")
def serve_menu_image(filename):
    return send_from_directory("static/menu", filename)

@app.route("/manager/menu/upload", methods=["POST"])
def add_menu_with_image():
    try:
        name = request.form.get("name")
        description = request.form.get("description", "")
        price = request.form.get("price")
        category = request.form.get("category", "main")
        image = request.files.get("image")

        if not name or not price or not image:
            return jsonify({"error": "Missing fields"}), 400

        if not allowed_file(image.filename):
            return jsonify({"error": "Invalid image format"}), 400

        filename = secure_filename(image.filename)
        image.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))

        image_path = f"/static/menu/{filename}"

        # ✅ PostgreSQL connection
        conn = get_db_connection()
        cur = conn.cursor()

        # ✅ FIXED query (%s instead of %s)
        cur.execute("""
            INSERT INTO menu_items (name, description, price, image, category)
            VALUES (%s, %s, %s, %s, %s)
        """, (name, description, price, image_path, category))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({"message": "Menu item added successfully"}), 201

    except Exception as e:
        print("UPLOAD ERROR:", e)
        return jsonify({"error": "Server error"}), 500

# ---------------------------
# Auth: Register & Login
# ---------------------------
@app.route("/register", methods=["POST"])
def register_user():
    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip()
    phone = (data.get("phone") or "").strip()
    password = (data.get("password") or "").strip()

    if not name or not email or not phone or not password:
        return jsonify({"error": "All fields are required"}), 400

    conn = get_db_connection()
    cur = conn.cursor()

    # Check duplicates
    cur.execute("SELECT id FROM customers WHERE email = %s", (email,))
    if cur.fetchone():
        conn.close()
        return jsonify({"error": "Email already registered"}), 400

    cur.execute("INSERT INTO customers (name, email, phone, password) VALUES (%s, %s, %s, %s)",
                (name, email, phone, password))
    conn.commit()
    conn.close()

    return jsonify({"message": "Registration successful!"})


@app.route("/login", methods=["POST"])
def universal_login():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip()
    password = (data.get("password") or "").strip()

    if not email or not password:
        return jsonify({"error": "Email & password required"}), 400

    conn = get_db_connection()
    cur = conn.cursor()

    # 1) check chefs
    cur.execute("SELECT id, name, email, password, active FROM chefs WHERE email = %s", (email,))
    chef = cur.fetchone()
    if chef:
        if chef["active"] == 0:
            conn.close()
            return jsonify({"error": "Your account is deactivated by owner"}), 403

        if password == chef["password"]:
            conn.close()
            return jsonify({
                "message": "Login successful!",
                "user": {"id": chef["id"], "name": chef["name"], "email": chef["email"], "user_type": "chef"}
            }), 200
        else:
            conn.close()
            return jsonify({"error": "Wrong password"}), 401
   
   # 1) check waiters
    cur.execute("SELECT id, name, email, password, active FROM waiters WHERE email = %s",(email,))
    waiter = cur.fetchone()
    if waiter:
            if waiter["active"] == 0:
                conn.close()
                return jsonify({"error": "Your account is deactivated by owner"}), 403
            if password == waiter["password"]:
                conn.close()
                return jsonify({
                    "message": "Login successful!",
                    "user": {"id": waiter["id"], "name": waiter["name"], "email": waiter["email"], "user_type": "waiter"}  # ✅ VERY IMPORTANT
                }), 200
            else:
                conn.close()
                return jsonify({"error": "Wrong password"}), 401

    # 2) check customers
    cur.execute("SELECT id, name, email, password FROM customers WHERE email = %s", (email,))
    customer = cur.fetchone()
    if customer:
        if password == customer["password"]:
            conn.close()
            return jsonify({
                "message": "Login successful!",
                "user": {"id": customer["id"], "name": customer["name"], "email": customer["email"], "user_type": "customer"}
            }), 200
        else:
            conn.close()
            return jsonify({"error": "Wrong password"}), 401

    # ✅ check managers
    cur.execute("SELECT id, name, email, password, active FROM managers WHERE email = %s", (email,))
    manager = cur.fetchone()

    if manager:
        if manager["active"] == 0:
            conn.close()
            return jsonify({"error": "Your account is deactivated by owner"}), 403

        if password == manager["password"]:
            conn.close()
            return jsonify({
                "message": "Login successful!",
                "user": {
                    "id": manager["id"],
                    "name": manager["name"],
                    "email": manager["email"],
                    "user_type": "manager"
                }
            }), 200

        else:
            conn.close()
            return jsonify({"error": "Wrong password"}), 401
    
    # ✅ check owners
    cur.execute("SELECT id, name, email, password FROM owners WHERE email = %s", (email,))
    owner = cur.fetchone()

    if owner:
        if password == owner["password"]:
            conn.close()
            return jsonify({
                "message": "Login successful!",
                "user": {
                    "id": owner["id"],
                    "name": owner["name"],
                    "email": owner["email"],
                    "user_type": "owner"
                }
            }), 200
        else:
            conn.close()
            return jsonify({"error": "Wrong password"}), 401

    conn.close()
    return jsonify({"error": "User not found"}), 404


# ---------------------------
# Menu
# ---------------------------
@app.route("/menu", methods=["GET"])
def get_menu():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT 
                id,
                name,
                description,
                price,
                image,
                IFNULL(category, 'main') AS category,
                IFNULL(is_available, 1) AS is_available
            FROM menu_items
            WHERE is_available = 1
              AND is_deleted = 0
            ORDER BY id DESC
        """)

        rows = cur.fetchall()
        conn.close()

        return jsonify([dict(r) for r in rows]), 200

    except Exception as e:
        print("Menu error:", e)
        return jsonify({"error": "Server error"}), 500

# ---------------------------
# Place Order (clean)
# ---------------------------
@app.route("/orders", methods=["POST"])
def place_order():
    try:
        data = request.get_json() or {}

        user_id = data.get("user_id") or data.get("userId")
        table_no = data.get("table_no") or data.get("tableNo")
        items = data.get("items", [])

        if not user_id or not table_no or not items:
            return jsonify({"success": False, "error": "Invalid request"}), 400

        conn = get_db_connection()
        cur = conn.cursor()

        # 1️⃣ Validate table
        cur.execute("""
            SELECT id, active FROM tables
            WHERE table_number = %s
        """, (table_no,))
        table = cur.fetchone()

        if not table:
            cur.close()
            conn.close()
            return jsonify({"success": False, "error": "Invalid table"}), 400

        table_id = table["id"]

        # 2️⃣ Active session
        cur.execute("""
            SELECT id FROM table_sessions
            WHERE table_no = %s AND status = 'active'
        """, (table_no,))
        session = cur.fetchone()

        if session:
            session_id = session["id"]
        else:
            started_at = datetime.now()

            cur.execute("""
                INSERT INTO table_sessions (table_no, started_at, status)
                VALUES (%s, %s, 'active')
                RETURNING id
            """, (table_no, started_at))

            session_id = cur.fetchone()["id"]

            cur.execute("""
                UPDATE tables SET active = 0 WHERE id = %s
            """, (table_id,))

        # 3️⃣ Calculate total
        total = sum(float(i["price"]) * int(i["quantity"]) for i in items)
        created_at = datetime.now()

        # 4️⃣ Create order
        cur.execute("""
            INSERT INTO orders
            (user_id, table_id, table_no, session_id, total, paid, status, created_at)
            VALUES (%s, %s, %s, %s, %s, 0, 'pending', %s)
            RETURNING id
        """, (user_id, table_id, table_no, session_id, total, created_at))

        order_id = cur.fetchone()["id"]

        # 5️⃣ Snapshot items
        socket_items = []

        for i in items:
            cur.execute("""
                SELECT name, image, price
                FROM menu_items
                WHERE id = %s
            """, (i["menu_id"],))

            menu = cur.fetchone()

            if menu:
                item_name = menu["name"]
                item_image = menu["image"] or ""
                price = menu["price"]
            else:
                item_name = "Deleted Item"
                item_image = ""
                price = float(i["price"])

            cur.execute("""
                INSERT INTO order_items
                (order_id, menu_id, item_name, item_image, quantity, price)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                order_id,
                i["menu_id"],
                item_name,
                item_image,
                i["quantity"],
                price
            ))

            socket_items.append({
                "name": item_name,
                "quantity": i["quantity"],
                "price": price
            })

        conn.commit()
        cur.close()
        conn.close()

        # 6️⃣ Realtime update
        socketio.emit(
            "order_created",
            {
                "order_id": order_id,
                "table_no": table_no,
                "session_id": session_id,
                "status": "pending",
                "items": socket_items
            },
            room="chef"
        )

        return jsonify({
            "success": True,
            "orderId": order_id,
            "sessionId": session_id
        }), 201

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/order/summary", methods=["POST"])
def order_summary():
    data = request.json
    items = data.get("items", [])

    subtotal = sum(i["price"] * i["quantity"] for i in items)

    gst = round(subtotal * 0.05, 2)
    service = round(subtotal * 0.05, 2)
    total = round(subtotal + gst + service, 2)

    return jsonify({
        "subtotal": subtotal,
        "gst": gst,
        "service": service,
        "total": total
    })

# ===============================
# CUSTOMER – GET ALL ORDERS
# ===============================
@app.route("/orders/user/<int:user_id>", methods=["GET"])
def get_customer_orders(user_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT 
                id,
                table_no,
                total,
                paid,
                status,
                created_at
            FROM orders
            WHERE user_id = %s
            ORDER BY created_at DESC
        """, (user_id,))

        orders = cur.fetchall()
        result = []

        for o in orders:
            cur.execute("""
                SELECT 
                    item_name AS name,
                    price,
                    quantity
                FROM order_items
                WHERE order_id = %s
            """, (o["id"],))
            items = [dict(i) for i in cur.fetchall()]

            result.append({
                "order_id": o["id"],
                "table_no": o["table_no"],
                "total": o["total"],
                "paid": bool(o["paid"]),
                "status": o["status"],
                "created_at": o["created_at"],
                "items": items
            })

        conn.close()
        return jsonify(result), 200

    except Exception as e:
        print("Customer orders error:", e)
        return jsonify([]), 200

# Fetch orders for a given user
@app.route("/orders/<int:order_id>", methods=["GET"])
def get_order(order_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, table_no, total, status, created_at
        FROM orders
        WHERE id = %s
    """, (order_id,))
    order = cur.fetchone()

    if not order:
        conn.close()
        return jsonify({"success": False}), 404

    cur.execute("""
        SELECT 
            item_name AS name,
            item_image AS image,
            quantity,
            price
        FROM order_items
        WHERE order_id = %s
    """, (order_id,))

    items = [dict(row) for row in cur.fetchall()]
    conn.close()

    return jsonify({
        "success": True,
        "order": dict(order),
        "items": items
    })

#-----------------------------------
#Bills
#-----------------------------------
from datetime import datetime

@app.route("/config/tax", methods=["GET"])
def get_tax_config():
    return jsonify(TAX_CONFIG), 200

@app.route("/bills/session/<int:table_no>", methods=["GET"])
def get_bill_by_table_session(table_no):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # 🔹 Active session
        cur.execute("""
            SELECT id FROM table_sessions
            WHERE table_no = %s AND status = 'active'
        """, (table_no,))
        session = cur.fetchone()

        if not session:
            return jsonify({
                "items": [],
                "subtotal": 0,
                "gst": 0,
                "service": 0,
                "total": 0,
                "paid": True
            })

        session_id = session["id"]

        # 🔹 Fetch items
        cur.execute("""
            SELECT
                oi.item_name AS name,
                SUM(oi.quantity) AS quantity,
                oi.price,
                SUM(oi.quantity * oi.price) AS subtotal
            FROM orders o
            JOIN order_items oi ON o.id = oi.order_id
            WHERE o.session_id = %s AND o.paid = 0
            GROUP BY oi.item_name, oi.price
        """, (session_id,))

        items = cur.fetchall()  # already dict (RealDictCursor)

        subtotal = sum(i["subtotal"] for i in items)

        gst = subtotal * TAX_CONFIG["gst"]
        service = subtotal * TAX_CONFIG["service"]
        total = subtotal + gst + service

        return jsonify({
            "session_id": session_id,
            "items": items,
            "subtotal": round(subtotal, 2),
            "gst": round(gst, 2),
            "service": round(service, 2),
            "total": round(total, 2),
            "paid": False
        })

    finally:
        cur.close()
        conn.close()

@app.route("/pay/session/<int:table_no>", methods=["POST"])
def pay_session(table_no):
    data = request.get_json()
    method = data.get("method", "Cash")

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔎 Get active session
    cur.execute("""
        SELECT id FROM table_sessions
        WHERE table_no = %s AND status = 'active'
    """, (table_no,))
    session = cur.fetchone()

    if not session:
        conn.close()
        return jsonify({"error": "No active session"}), 404

    session_id = session["id"]

    # 🔎 Get all unpaid orders in session
    cur.execute("""
        SELECT o.id FROM orders o
        WHERE o.session_id = %s AND paid = 0
    """, (session_id,))
    unpaid_orders = [r["id"] for r in cur.fetchall()]

    if not unpaid_orders:
        conn.close()
        return jsonify({"error": "No unpaid orders"}), 400

    # 🔎 Calculate subtotal, GST, Service, total
    cur.execute("""
        SELECT SUM(oi.quantity * oi.price) AS subtotal
        FROM order_items oi
        JOIN orders o ON oi.order_id = o.id
        WHERE o.session_id = %s AND o.paid = 0
    """, (session_id,))
    subtotal = cur.fetchone()["subtotal"] or 0

    gst = subtotal * TAX_CONFIG["gst"]
    service = subtotal * TAX_CONFIG["service"]
    total_final = subtotal + gst + service

    # 🔎 Update orders: mark as paid, completed, store total
    cur.execute(f"""
        UPDATE orders
        SET paid = 1,
            status = 'completed',
            payment_method = %s,
            total = %s
        WHERE session_id = %s
    """, (method, round(total_final, 2), session_id))

    # 🔎 Close session
    cur.execute("""
        UPDATE table_sessions
        SET status = 'closed',
            ended_at = CURRENT_TIMESTAMP
        WHERE id = %s
    """, (session_id,))

    # 🔎 Free table
    cur.execute("""
        UPDATE tables
        SET active = 0
        WHERE table_number = %s
    """, (table_no,))

    conn.commit()
    conn.close()

    return jsonify({"success": True, "method": method, "total": round(total_final, 2)}), 200

# @app.route("/bills/history/<int:table_no>")
# def bill_history(table_no):
#     conn = get_db_connection()
#     conn.row_factory = sqlite3.Row
#     cur = conn.cursor()

#     cur.execute("""
#         SELECT
#             ts.id AS session_id,
#             ts.ended_at,
#             SUM(oi.quantity * oi.price) AS total
#         FROM table_sessions ts
#         JOIN orders o ON ts.id = o.session_id
#         JOIN order_items oi ON o.id = oi.order_id
#         WHERE ts.table_no = %s
#           AND ts.status = 'closed'
#           AND o.paid = 1
#         GROUP BY ts.id
#         ORDER BY ts.ended_at DESC
#     """, (table_no,))

#     return jsonify([dict(row) for row in cur.fetchall()]), 200

@app.route("/bills/history/user/<int:user_id>", methods=["GET"])
def bill_history_by_user(user_id):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT 
                o.id AS order_id,
                o.table_no,
                o.session_id,
                o.total,
                o.payment_method,
                o.created_at
            FROM orders o
            WHERE o.user_id = %s
              AND o.paid = 1
            ORDER BY o.created_at DESC
        """, (user_id,))

        orders = cur.fetchall()
        result = []

        for o in orders:
            cur.execute("""
                SELECT 
                    item_name AS name,
                    quantity,
                    price,
                    (quantity * price) AS subtotal
                FROM order_items
                WHERE order_id = %s
            """, (o["order_id"],))

            items = cur.fetchall()  # already dict

            result.append({
                "order_id": o["order_id"],
                "table_no": o["table_no"],
                "session_id": o["session_id"],
                "items": items,
                "total": o["total"],
                "payment_method": o["payment_method"],
                "date": o["created_at"]
            })

        return jsonify(result), 200

    except Exception as e:
        print("Bill history error:", e)
        return jsonify({"error": "Server error"}), 500

    finally:
        cur.close()
        conn.close()
        
from datetime import datetime

@app.route("/bills/invoice/<int:session_id>")
def get_invoice(session_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # get orders for session
        cur.execute("""
            SELECT o.id, o.payment_method, o.created_at
            FROM orders o
            WHERE o.session_id = %s
        """, (session_id,))
        order = cur.fetchone()

        if not order:
            return jsonify({"error": "Invoice not found"}), 404

        # get items
        cur.execute("""
            SELECT
                oi.item_name AS name,
                oi.quantity,
                oi.price,
                (oi.quantity * oi.price) AS subtotal
            FROM order_items oi
            JOIN orders o ON oi.order_id = o.id
            WHERE o.session_id = %s
        """, (session_id,))

        items = [dict(row) for row in cur.fetchall()]

        subtotal = sum(i["subtotal"] for i in items)
        gst = round(subtotal * TAX_CONFIG["gst"], 2)
        service = round(subtotal * TAX_CONFIG["service"], 2)
        total = round(subtotal + gst + service, 2)

        return jsonify({
            "invoice_number": f"INV-{session_id:05d}",
            "session_id": session_id,
            "items": items,
            "subtotal": subtotal,
            "gst": gst,
            "service": service,
            "total": total,
            "payment_method": order["payment_method"],
            "created_at": order["created_at"]
        }),200

    except Exception as e:
        print("INVOICE ERROR:", e)
        return jsonify({"error": "Invoice generation failed"}), 500

    finally:
        conn.close()

@app.route("/orders/unpaid/<int:table_no>")
def get_unpaid_orders(table_no):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # 🔎 get active session
        cur.execute("""
            SELECT id
            FROM table_sessions
            WHERE table_no = %s AND status = 'active'
        """, (table_no,))
        session = cur.fetchone()

        if not session:
            return jsonify([]), 200

        session_id = session["id"]

        # 🔎 unpaid orders
        cur.execute("""
            SELECT o.id, o.created_at, o.status, o.total
            FROM orders o
            WHERE o.session_id = %s AND o.paid = 0
            ORDER BY o.created_at DESC
        """, (session_id,))

        orders = cur.fetchall()  # ✅ already dict
        return jsonify(orders), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        cur.close()
        conn.close()

#----------------------------
#Payment 
#-----------------------------
@app.route("/create-checkout-session", methods=["POST"])
def create_checkout_session():
    try:
        data = request.json
        table_no = data["table_no"]
        amount = int(float(data["amount"]) * 100)  # convert to paisa

        session = stripe.checkout.Session.create(
            mode="payment",
            line_items=[{
                "price_data": {
                    "currency": "inr",   # ✅ INR currency
                    "product_data": {
                        "name": f"Eatrova Table {table_no} Bill"
                    },
                    "unit_amount": amount,
                },
                "quantity": 1,
            }],
            success_url="http://localhost:5173/payment-success%ssession_id={CHECKOUT_SESSION_ID}",
            cancel_url="http://localhost:5173/payment-cancel",
            metadata={"table_no": table_no}
        )

        return jsonify({"url": session.url})

    except Exception as e:
        print("Stripe Error:", str(e))
        return jsonify({"error": str(e)}), 500

@app.route("/verify-payment", methods=["POST"])
def verify_payment():
    session_id = request.json.get("session_id")

    session = stripe.checkout.Session.retrieve(session_id)

    if session.payment_status == "paid":
        table_no = session.metadata["table_no"]
        amount = session.amount_total // 100

        conn = get_db_connection()
        cur = conn.cursor()

        # Insert payment record
        cur.execute("""
            INSERT INTO payments 
            (table_no, stripe_session_id, stripe_payment_intent, amount, method, status)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            table_no,
            session.id,
            session.payment_intent,
            amount,
            "Card",
            "Success"
        ))

        # Update table session
        cur.execute("""
            UPDATE table_sessions
            SET status='paid'
            WHERE table_no=%s AND status='active'
        """, (table_no,))

        conn.commit()
        conn.close()

        return jsonify({"message": "Payment Verified"})

    return jsonify({"message": "Payment Failed"})

# ---------------------------
# Dashboard helper endpoints (optional)
# ---------------------------
@app.route("/dashboard/unpaid_counts", methods=["GET"])
def dashboard_unpaid_counts():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT user_id, COUNT(*) as unpaid_count
            FROM orders
            WHERE paid = 0
            GROUP BY user_id
        """)
        rows = cur.fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows]), 200
    except Exception as e:
        print("dashboard_unpaid_counts error:", e)
        return jsonify({"error": "Server error"}), 500


# ------------------------------
# UPDATE ORDER STATUS
# ------------------------------
@app.route("/orders/<int:order_id>/status", methods=["PUT"])
def update_order_status(order_id):
    try:
        data = request.get_json(force=True)
        new_status = data.get("status")

        if not new_status:
            return jsonify({"error": "Status required"}), 400

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("SELECT id FROM orders WHERE id = %s", (order_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify({"error": "Order not found"}), 404

        cur.execute("""
            UPDATE orders
            SET status = %s, updated_at = %s
            WHERE id = %s
        """, (
            new_status,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            order_id
        ))

        conn.commit()
        conn.close()

        # 🔔 SOCKET UPDATE
        socketio.emit("order_updated", {
            "order_id": order_id,
            "status": new_status
        })

        return jsonify({"success": True}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Server error"}), 500

@app.route("/orders/details/<int:order_id>", methods=["GET"])
def get_order_details(order_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT id, table_no, total, status
            FROM orders
            WHERE id = %s
        """, (order_id,))
        order = cur.fetchone()

        if not order:
            conn.close()
            return jsonify({"success": False}), 404

        cur.execute("""
            SELECT 
                item_name AS name,
                item_image AS image,
                quantity,
                price
            FROM order_items
            WHERE order_id = %s
        """, (order_id,))

        items = [dict(row) for row in cur.fetchall()]
        conn.close()

        return jsonify({
            "success": True,
            "order": dict(order),
            "items": items
        })

    except Exception as e:
        print("ORDER FETCH ERROR:", e)
        return jsonify({"success": False, "error": str(e)}), 500

# ------------------------------
# CANCEL ORDER (NEW)
# ------------------------------
@app.route("/orders/<int:order_id>/cancel", methods=["PUT"])
def cancel_order(order_id):
    try:
        data = request.get_json() or {}
        reason = data.get("reason", "")

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("SELECT user_id FROM orders WHERE id = %s", (order_id,))
        row = cur.fetchone()

        if not row:
            conn.close()
            return jsonify({"error": "Order not found"}), 404

        user_id = row[0] if not isinstance(row, dict) else row["user_id"]

        cur.execute("""
            UPDATE orders
            SET status = 'cancelled',
                cancel_reason = %s,
                updated_at = %s
            WHERE id = %s
        """, (reason, timestamp, order_id))

        conn.commit()
        conn.close()

        payload = {
            "order_id": order_id,
            "status": "cancelled",
            "updated_at": timestamp,
            "user_id": user_id,
            "reason": reason
        }

        # Emit safely
        try:
            socketio.emit("order_cancelled", payload, room="chef")
            socketio.emit("order_cancelled", payload, room=f"user_{user_id}")
            socketio.emit("order_cancelled", payload)
        except Exception as e:
            print("Socket emit error:", e)

        print(f"Order #{order_id} cancelled")

        return jsonify({"message": "Order cancelled", "data": payload}), 200

    except Exception as e:
        print("Cancel Order Error:", e)
        return jsonify({"error": str(e)}), 500
    
# ------------------------------
# Update item status (existing)
# ------------------------------
@app.route("/orders/<int:order_id>/item/<int:item_id>/status", methods=["PUT"])
def update_item_status(order_id, item_id):
    try:
        data = request.get_json() or {}
        new_status = data.get("status")
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        conn = get_db_connection()
        cur = conn.cursor()

        if new_status == "preparing":
            cur.execute("""
                UPDATE order_items 
                SET started_at = %s 
                WHERE id = %s AND order_id = %s
            """, (ts, item_id, order_id))

        if new_status == "ready":
            cur.execute("""
                UPDATE order_items 
                SET finished_at = %s 
                WHERE id = %s AND order_id = %s
            """, (ts, item_id, order_id))

        conn.commit()
        conn.close()

        socketio.emit("order_item_updated", {
            "order_id": order_id,
            "item_id": item_id,
            "status": new_status,
            "timestamp": ts
        }, room="chef")
        socketio.emit("item_status_updated", {
            "order_id": order_id,
            "item_id": item_id,
            "status": new_status
        })

        log_event(f"Chef updated item #{item_id} in Order #{order_id}")

        return jsonify({"message": "Item updated"}), 200

    except Exception as e:
        print("Item status update error:", e)
        return jsonify({"error": str(e)}), 500

# ---------------------------
# Chef orders endpoint (returns current orders including cancelled)
@app.route("/chef/orders", methods=["GET"])
def chef_orders():
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                o.id AS order_id,
                o.table_no,
                o.total,
                o.status,
                o.created_at,
                o.cancel_reason,
                oi.item_name,
                oi.quantity
            FROM orders o
            LEFT JOIN order_items oi ON o.id = oi.order_id
            WHERE o.status IN ('pending','preparing','ready','cancelled')
            ORDER BY o.created_at DESC
        """)

        rows = cur.fetchall()

        orders = {}

        for r in rows:
            oid = r["order_id"]

            if oid not in orders:
                orders[oid] = {
                    "order_id": oid,
                    "table_no": r["table_no"],
                    "total": r["total"],
                    "status": r["status"],
                    "created_at": r["created_at"],
                    "cancel_reason": r["cancel_reason"] or "",
                    "items": []
                }

            if r["item_name"]:
                orders[oid]["items"].append({
                    "name": r["item_name"],
                    "quantity": r["quantity"]
                })

        return jsonify(list(orders.values())), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        cur.close()
        conn.close()
    
@app.route("/waiter/orders", methods=["GET"])
def waiter_orders():
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT id, table_no, total, status, created_at
            FROM orders
            WHERE status IN ('ready','completed','cancelled')
            ORDER BY created_at DESC
        """)
        orders_rows = cur.fetchall()

        orders = []

        for o in orders_rows:
            cur.execute("""
                SELECT item_name AS name, quantity
                FROM order_items
                WHERE order_id = %s
            """, (o["id"],))

            items = cur.fetchall()  # ✅ already dict

            orders.append({
                "order_id": o["id"],
                "table_no": o["table_no"],
                "total": o["total"],
                "status": o["status"],
                "created_at": o["created_at"],
                "items": items
            })

        return jsonify(orders), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        cur.close()
        conn.close()

@app.route("/orders", methods=["GET"])
def get_all_orders():
    conn = get_db_connection()
    orders = conn.execute("SELECT * FROM orders ORDER BY created_at DESC").fetchall()
    conn.close()

    orders_list = []
    for order in orders:
        orders_list.append(dict(order))

    return jsonify(orders_list), 200

# ------------------------------
# SERVE ORDER (FOR WAITER)
# ------------------------------
@app.route("/order/<int:order_id>/serve", methods=["PUT"])
def serve_order(order_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # 1️⃣ Check order exists
        cur.execute("""
            SELECT status, table_no, session_id
            FROM orders
            WHERE id = %s
        """, (order_id,))
        order = cur.fetchone()

        if not order:
            cur.close()
            conn.close()
            return jsonify({"success": False, "error": "Order not found"}), 404

        # 2️⃣ Only READY orders can be served
        if order["status"] != "ready":
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "error": f"Cannot serve order with status '{order['status']}'"
            }), 400

        # 3️⃣ Update order status
        cur.execute("""
            UPDATE orders
            SET status = 'completed',
                updated_at = %s
            WHERE id = %s
        """, (
            datetime.now(),
            order_id
        ))

        # 4️⃣ Close table session
        if order["session_id"]:
            cur.execute("""
                UPDATE table_sessions
                SET status = 'closed', ended_at = %s
                WHERE id = %s
            """, (
                datetime.now(),
                order["session_id"]
            ))

        conn.commit()

        # ✅ Close properly
        cur.close()
        conn.close()

        # 5️⃣ Notify all dashboards
        socketio.emit("order_updated", {"order_id": order_id})

        return jsonify({"success": True}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": "Server error"
        }), 500
    
@app.route("/manager/orders", methods=["GET"])
def manager_all_orders():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT 
                o.id,
                o.user_id,
                c.name AS customer_name,
                o.table_no,
                o.total,
                o.paid,
                o.status,
                o.created_at,
                IFNULL(o.cancel_reason,'') AS cancel_reason
            FROM orders o
            LEFT JOIN customers c ON c.id = o.user_id
            ORDER BY o.created_at DESC
        """)

        orders = cur.fetchall()
        result = []

        for o in orders:
            cur.execute("""
                SELECT 
                    item_name,
                    item_image,
                    quantity,
                    price,
                    (quantity * price) AS subtotal
                FROM order_items
                WHERE order_id = %s
            """, (o["id"],))

            items = [dict(i) for i in cur.fetchall()]

            result.append({
                "id": o["id"],
                "user_id": o["user_id"],
                "customer_name": o["customer_name"],
                "tableNumber": o["table_no"],
                "totalAmount": o["total"],
                "paid": bool(o["paid"]),
                "status": o["status"],
                "createdAt": o["created_at"],
                "cancelReason": o["cancel_reason"],
                "items": items
            })

        conn.close()
        return jsonify(result), 200

    except Exception as e:
        print("Manager orders error:", e)
        return jsonify({"error": str(e)}), 500

#--------------------------------------
import uuid

ALLOWED_EXTENSIONS = {"png","jpg","jpeg","webp"}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".",1)[1].lower() in ALLOWED_EXTENSIONS

# ---------------------------
# Manager menu - READ all
# ---------------------------
@app.route("/manager/menu", methods=["GET"])
def manager_get_menu():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            id, name, description, price, image,
            IFNULL(category,'main') AS category,
            is_available,
            is_deleted,
            deleted_at
        FROM menu_items
        ORDER BY id DESC
    """)

    rows = cur.fetchall()
    conn.close()

    return jsonify([dict(r) for r in rows]), 200

# ---------------------------
# Manager menu - CREATE
# (If you already have /manager/menu POST, replace it with this one to emit events)
# ---------------------------
@app.route("/manager/menu", methods=["POST"])
def manager_add_menu():
    try:
        name = request.form.get("name")
        description = request.form.get("description","")
        price = request.form.get("price")
        category = request.form.get("category","main")

        if not name or price is None:
            return jsonify({"error":"Name & price required"}), 400

        price = float(price)
        image_path = ""

        file = request.files.get("image")
        if file and file.filename:
            if not allowed_file(file.filename):
                return jsonify({"error":"Invalid image type"}), 400

            ext = file.filename.rsplit(".",1)[1].lower()
            filename = f"{uuid.uuid4().hex}.{ext}"

            os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
            file.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))
            image_path = f"/static/menu/{filename}"

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO menu_items
            (name, description, price, image, category, is_available, is_deleted)
            VALUES (%s, %s, %s, %s, %s, 1, 0)
        """, (name, description, price, image_path, category))

        conn.commit()
        item_id = cur.lastrowid
        conn.close()

        socketio.emit("menu_changed", {"action":"created","item_id":item_id})

        return jsonify({"id": item_id}), 201

    except Exception as e:
        print("manager_add_menu error:", e)
        return jsonify({"error":"Server error"}), 500

# ---------------------------
# Manager menu - DELETE
# ---------------------------
@app.route("/manager/menu/<int:item_id>/delete", methods=["POST"])
def delete_menu_item(item_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE menu_items
        SET is_deleted = 1,
            is_available = 0,
            deleted_at = datetime('now')
        WHERE id = %s
    """, (item_id,))

    conn.commit()
    conn.close()

    socketio.emit("menu_changed", {"action":"deleted","item_id":item_id})

    return jsonify({"success": True})

# ---------------------------
# Manager menu - UPDATE
# ---------------------------
@app.route("/manager/menu/<int:item_id>", methods=["PUT"])
def manager_update_menu(item_id):
    try:
        name = request.form.get("name")
        description = request.form.get("description")
        price = request.form.get("price")
        category = request.form.get("category")

        fields, values = [], []

        if name:
            fields.append("name=%s")
            values.append(name)
        if description is not None:
            fields.append("description=%s")
            values.append(description)
        if price:
            fields.append("price=%s")
            values.append(float(price))
        if category:
            fields.append("category=%s")
            values.append(category)

        file = request.files.get("image")
        if file and file.filename:
            ext = file.filename.rsplit(".",1)[1].lower()
            filename = f"{uuid.uuid4().hex}.{ext}"
            file.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))
            fields.append("image=%s")
            values.append(f"/static/menu/{filename}")

        if not fields:
            return jsonify({"error":"No data"}), 400

        values.append(item_id)

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(f"""
            UPDATE menu_items SET {", ".join(fields)}
            WHERE id=%s
        """, values)
        conn.commit()
        conn.close()

        socketio.emit("menu_changed", {"action":"updated","item_id":item_id})
        log_event(f"Manager updated menu item #{item_id}")

        return jsonify({"success":True}), 200

    except Exception as e:
        print("update error:", e)
        return jsonify({"error":"Server error"}), 500

#-----------------------------------
# ✅ TOGGLE MENU ITEM AVAILABILITY
#-----------------------------------
@app.route("/manager/menu/<int:item_id>/toggle", methods=["POST"])
def toggle_menu_item(item_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE menu_items
        SET is_available = CASE 
            WHEN is_available = 1 THEN 0
            ELSE 1
        END
        WHERE id = %s AND is_deleted = 0
    """, (item_id,))

    conn.commit()

    cur.execute("""
        SELECT is_available
        FROM menu_items
        WHERE id = %s
    """, (item_id,))
    row = cur.fetchone()

    conn.close()

    return jsonify({
        "success": True,
        "available": bool(row["is_available"])
    })

@app.route("/manager/menu/deleted", methods=["GET"])
def get_deleted_menu():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT * FROM menu_items
        WHERE is_deleted = 1
        ORDER BY deleted_at DESC
    """)

    items = cur.fetchall()
    conn.close()

    return jsonify([dict(row) for row in items])

@app.route("/manager/menu/<int:item_id>/restore", methods=["PATCH"])
def restore_menu_item(item_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE menu_items
        SET is_deleted = 0,
            is_available = 1,
            deleted_at = NULL
        WHERE id = %s
    """, (item_id,))

    conn.commit()
    conn.close()

    socketio.emit("menu_changed", {"action":"restored","item_id":item_id})

    return jsonify({"success": True})

# ----------------------------------------------
# Inventory
# ----------------------------------------------

# GET all inventory items + alerts
@app.route("/inventory", methods=["GET"])
def get_inventory():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM inventory")
    items = cur.fetchall()

    inventory = []
    alerts = {
        "low": [],
        "out": []
    }

    for item in items:
        if item["current_stock"] <= 0:
            status = "out"
            alerts["out"].append(item["name"])
        elif item["current_stock"] <= item["min_stock"]:
            status = "low"
            alerts["low"].append(item["name"])
        else:
            status = "well"

        inventory.append({
            "id": item["id"],
            "name": item["name"],
            "current_stock": item["current_stock"],
            "min_stock": item["min_stock"],
            "unit": item["unit"],
            "status": status
        })

    conn.close()

    return jsonify({
        "items": inventory,
        "alerts": alerts
    })


# ADD new inventory item
@app.route("/inventory", methods=["POST"])
def create_inventory_item():
    data = request.json

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO inventory (name, current_stock, min_stock, unit)
        VALUES (%s, %s, %s, %s)
    """, (
        data["name"],
        data.get("current_stock", 0),
        data["min_stock"],
        data.get("unit", "kg")
    ))

    conn.commit()
    conn.close()

    return jsonify({"success": True})


# UPDATE stock only (+ / -)
@app.route("/inventory/<int:item_id>/stock", methods=["PATCH"])
def update_inventory_stock(item_id):
    data = request.json

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE inventory
        SET current_stock = MAX(current_stock + %s, 0)
        WHERE id = %s
    """, (data["change"], item_id))

    conn.commit()
    conn.close()

    return jsonify({"success": True})


# EDIT item details (name / min / unit)
@app.route("/inventory/<int:item_id>", methods=["PUT"])
def edit_inventory_item(item_id):
    data = request.json

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE inventory
        SET name = %s, min_stock = %s, unit = %s
        WHERE id = %s
    """, (
        data["name"],
        data["min_stock"],
        data["unit"],
        item_id
    ))

    conn.commit()
    conn.close()

    return jsonify({"success": True})


# DELETE inventory item
@app.route("/inventory/<int:item_id>", methods=["DELETE"])
def delete_inventory_item(item_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM inventory WHERE id = %s", (item_id,))
    conn.commit()
    conn.close()

    return jsonify({"success": True})

#=============================================================

# -------------------- OWNER POWER ENDPOINTS (COPY / PASTE into app.py) --------------------
from flask import request, jsonify, Response
from datetime import datetime
import csv
import io

# NOTE: this block assumes you have these helpers defined earlier in app.py:
# - get_db_connection()  -> returns sqlite3 connection with row_factory producing dict-like rows
# - socketio (flask-socketio) initialized as socketio

# Helper to safely read numeric column names that may vary across DB versions
def _num(row, *keys, default=0):
    for k in keys:
        if k in row and row[k] is not None:
            try:
                return float(row[k])
            except Exception:
                pass
    return default

#-----------------------------------------
# Staff Management
#-----------------------------------------

# 1) Get all staff (managers, chefs, waiters) - normalized
@app.route("/owner/staff", methods=["GET"])
def owner_get_staff():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        staff = []

        cur.execute("SELECT id, name, email, active FROM managers")
        for r in cur.fetchall():
            row = dict(r); row["role"] = "manager"; staff.append(row)

        cur.execute("SELECT id, name, email, active FROM chefs")
        for r in cur.fetchall():
            row = dict(r); row["role"] = "chef"; staff.append(row)

        cur.execute("SELECT id, name, email, active FROM waiters")
        for r in cur.fetchall():
            row = dict(r); row["role"] = "waiter"; staff.append(row)

        conn.close()
        return jsonify(staff), 200
    except Exception as e:
        print("owner_get_staff error:", e)
        return jsonify({"error": "Server error"}), 500

# 2) Create staff (any role)
@app.route("/owner/staff/<role>", methods=["POST"])
def owner_create_staff(role):
    if role not in ["managers", "chefs", "waiters"]:
        return jsonify({"error": "Invalid role"}), 400
    try:
        data = request.get_json() or {}
        name = (data.get("name") or "").strip()
        email = (data.get("email") or "").strip()
        password = (data.get("password") or "").strip()
        if not name or not email or not password:
            return jsonify({"error": "Missing fields"}), 400

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(f"INSERT INTO {role} (name, email, password, active) VALUES (%s, %s, %s, 1)", (name, email, password))
        conn.commit()
        new_id = cur.lastrowid
        conn.close()

        try:
            socketio.emit("staff_changed", {"role": role, "id": new_id, "active": True})
        except Exception:
            pass

        return jsonify({"message": f"{role[:-1].capitalize()} created", "id": new_id}), 201
    except Exception as e:
        print("owner_create_staff error:", e)
        return jsonify({"error": "Server error"}), 500

# 3) Toggle staff active
@app.route("/owner/staff/<role>/<int:user_id>/toggle", methods=["PATCH"])
def owner_toggle_staff(role, user_id):
    if role not in ["managers", "chefs", "waiters"]:
        return jsonify({"error": "Invalid role"}), 400
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(f"SELECT active FROM {role} WHERE id = %s", (user_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "User not found"}), 404
        current = int(row["active"]) if row["active"] is not None else 1
        new_status = 0 if current == 1 else 1
        cur.execute(f"UPDATE {role} SET active = %s WHERE id = %s", (new_status, user_id))
        conn.commit()
        conn.close()
        try:
            socketio.emit("staff_changed", {"role": role, "id": user_id, "active": bool(new_status)})
        except Exception:
            pass
        return jsonify({"success": True, "active": bool(new_status)}), 200
    except Exception as e:
        print("owner_toggle_staff error:", e)
        return jsonify({"error": "Server error"}), 500

#-------------------------------------
# Restaurant Status
#-------------------------------------

# 4) Toggle restaurant open/close + status getter
@app.route("/owner/restaurant/toggle", methods=["PATCH"])
def owner_toggle_restaurant():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        # Ensure a row exists (id=1)
        cur.execute("SELECT is_open FROM restaurant_status WHERE id = 1")
        r = cur.fetchone()
        current = int(r["is_open"]) if r and r["is_open"] is not None else 1
        new_status = 0 if current == 1 else 1
        cur.execute("UPDATE restaurant_status SET is_open = %s WHERE id = 1", (new_status,))
        conn.commit()
        conn.close()
        try:
            socketio.emit("restaurant_status", {"open": bool(new_status)})
        except Exception:
            pass
        return jsonify({"open": bool(new_status)}), 200
    except Exception as e:
        print("owner_toggle_restaurant error:", e)
        return jsonify({"error": "Server error"}), 500

@app.route("/owner/restaurant/status", methods=["GET"])
def owner_get_restaurant_status():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT is_open FROM restaurant_status WHERE id = 1")
        r = cur.fetchone()
        conn.close()
        open_flag = bool(int(r["is_open"])) if r and r["is_open"] is not None else True
        return jsonify({"open": open_flag}), 200
    except Exception as e:
        print("owner_get_restaurant_status error:", e)
        return jsonify({"error": "Server error"}), 500

#------------------------------------
# Menu Control
#------------------------------------

# 5) Owner delete menu item
@app.route("/owner/menu/<int:item_id>", methods=["DELETE"])
def owner_delete_menu(item_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM menu_items WHERE id = %s", (item_id,))
        conn.commit()
        conn.close()
        try:
            socketio.emit("menu_deleted", {"id": item_id})
            socketio.emit("menu_changed", {"action": "deleted", "item_id": item_id})
        except Exception:
            pass
        return jsonify({"success": True}), 200
    except Exception as e:
        print("owner_delete_menu error:", e)
        return jsonify({"error": "Server error"}), 500

# 6) Toggle menu availability
@app.route("/owner/menu/<int:item_id>/toggle", methods=["PATCH"])
def owner_toggle_menu(item_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT available FROM menu_items WHERE id = %s", (item_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Item not found"}), 404
        current = int(row["available"]) if row["available"] is not None else 1
        new_status = 0 if current == 1 else 1
        cur.execute("UPDATE menu_items SET available = %s WHERE id = %s", (new_status, item_id))
        conn.commit()
        conn.close()
        try:
            socketio.emit("menu_changed", {"item_id": item_id, "available": bool(new_status)})
        except Exception:
            pass
        
        log_event(f"Menu item #{item_id} toggled by Owner")

        return jsonify({"success": True, "available": bool(new_status)}), 200
    except Exception as e:
        print("owner_toggle_menu error:", e)
        return jsonify({"error": "Server error"}), 500

#--------------------------------------
# Orders
#-------------------------------------

# 7) Block order (owner)
@app.route("/owner/order/<int:order_id>/block", methods=["PUT"])
def owner_block_order(order_id):
    try:
        data = request.get_json() or {}
        reason = (data.get("reason") or "Blocked by owner").strip()
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT user_id, status FROM orders WHERE id = %s", (order_id,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "Order not found"}), 404
        if row.get("status") == "completed":
            conn.close()
            return jsonify({"error": "Completed order can't be blocked"}), 400

        user_id = row.get("user_id")
        cur.execute("UPDATE orders SET status = 'cancelled', cancel_reason = %s, updated_at = %s WHERE id = %s", (reason, ts, order_id))
        conn.commit()
        conn.close()

        payload = {"order_id": order_id, "status": "cancelled", "updated_at": ts, "user_id": user_id, "reason": reason}
        try:
            socketio.emit("order_cancelled", payload, room="chef")
            socketio.emit("order_cancelled", payload, room=f"user_{user_id}")
            socketio.emit("order_cancelled", payload)
        except Exception:
            pass

        log_event(f"Owner blocked Order #{order_id}")

        return jsonify({"message": "Order blocked", "data": payload}), 200
    except Exception as e:
        print("owner_block_order error:", e)
        return jsonify({"error": "Server error"}), 500

#--------------------------------
# Reports
#---------------------------------

@app.route("/owner/report/daily")
def daily_revenue_report():

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT DATE(created_at) as date,
               COUNT(id) as orders,
               SUM(total) as revenue
        FROM orders
        GROUP BY DATE(created_at)
        ORDER BY DATE(created_at) DESC
        LIMIT 10
    """)

    rows = cursor.fetchall()

    report = []

    for r in rows:
        report.append({
            "date": r["date"],
            "orders": r["orders"],
            "revenue": r["revenue"]
        })

    conn.close()

    return jsonify(report)

@app.route("/owner/report/weekly")
def weekly_sales_report():

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT strftime('%W', created_at) as week,
               COUNT(id) as orders,
               SUM(total) as revenue
        FROM orders
        GROUP BY week
        ORDER BY week DESC
        LIMIT 8
    """)

    rows = cursor.fetchall()

    report = []

    for r in rows:
        report.append({
            "week": r["week"],
            "orders": r["orders"],
            "revenue": r["revenue"]
        })

    conn.close()

    return jsonify(report)

@app.route("/owner/report/monthly")
def monthly_revenue_report():

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT strftime('%Y-%m', created_at) as month,
               COUNT(id) as orders,
               SUM(total) as revenue
        FROM orders
        GROUP BY month
        ORDER BY month DESC
        LIMIT 12
    """)

    rows = cursor.fetchall()

    report = []

    for r in rows:
        report.append({
            "month": r["month"],
            "orders": r["orders"],
            "revenue": r["revenue"]
        })

    conn.close()

    return jsonify(report)

@app.route("/owner/report/category")
def category_revenue():

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT m.category,
               SUM(oi.quantity) as items_sold,
               SUM(oi.price * oi.quantity) as revenue
        FROM order_items oi
        JOIN menu m ON oi.item_name = m.name
        GROUP BY m.category
        ORDER BY revenue DESC
    """)

    rows = cursor.fetchall()

    report = []

    for r in rows:
        report.append({
            "category": r["category"],
            "items_sold": r["items_sold"],
            "revenue": r["revenue"]
        })

    conn.close()

    return jsonify(report)

@app.route("/owner/report/orders")
def order_history():

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id,
            table_no,
            SUM(total) as amount,
            status,
            created_at
        FROM orders
        ORDER BY created_at DESC
        LIMIT 30
    """)

    rows = cursor.fetchall()

    report = []

    for r in rows:
        report.append({
            "order_id": r["id"],
            "table": r["table_no"],
            "amount": r["amount"],
            "status": r["status"],
            "date": r["created_at"]
        })

    conn.close()

    return jsonify(report)

@app.route("/owner/report/export")
def export_reports_excel():

    conn = get_db_connection()
    cursor = conn.cursor()

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

# ---------------- DAILY SHEET ----------------
    daily_sheet = wb.create_sheet("Daily Revenue")

    daily_sheet.append(["Date", "Orders", "Revenue"])

    cursor.execute("""
        SELECT DATE(created_at),
               COUNT(id),
               SUM(total)
        FROM orders
        GROUP BY DATE(created_at)
        ORDER BY DATE(created_at) DESC
    """)

    for row in cursor.fetchall():
        daily_sheet.append(tuple(row))

# ---------------- WEEKLY SHEET ----------------
    weekly_sheet = wb.create_sheet("Weekly Sales")

    weekly_sheet.append(["Week", "Orders", "Revenue"])

    cursor.execute("""
        SELECT strftime('%W', created_at),
               COUNT(id),
               SUM(total)
        FROM orders
        GROUP BY strftime('%W', created_at)
        ORDER BY strftime('%W', created_at) DESC
    """)

    for row in cursor.fetchall():
        weekly_sheet.append(tuple(row))

# ---------------- MONTHLY SHEET ----------------
    monthly_sheet = wb.create_sheet("Monthly Revenue")

    monthly_sheet.append(["Month", "Orders", "Revenue"])

    cursor.execute("""
        SELECT strftime('%Y-%m', created_at),
               COUNT(id),
               SUM(total)
        FROM orders
        GROUP BY strftime('%Y-%m', created_at)
        ORDER BY strftime('%Y-%m', created_at) DESC
    """)

    for row in cursor.fetchall():
        monthly_sheet.append(tuple(row))

# ---------------- CATEGORY SHEET ----------------
    category_sheet = wb.create_sheet("Category Revenue")

    category_sheet.append(["Category", "Items Sold", "Revenue"])

    cursor.execute("""
        SELECT m.category,
               SUM(oi.quantity),
               SUM(oi.price * oi.quantity)
        FROM order_items oi
        JOIN menu_items m ON oi.item_name = m.name
        GROUP BY m.category
        ORDER BY SUM(oi.price * oi.quantity) DESC
    """)

    for row in cursor.fetchall():
        category_sheet.append(tuple(row))

# ---------------- ORDER HISTORY SHEET ----------------
    orders_sheet = wb.create_sheet("Order History")

    orders_sheet.append(["Order ID", "Table", "Amount", "Status", "Date"])

    cursor.execute("""
        SELECT id,
            table_no,
            total,
            status,
            created_at
        FROM orders
        ORDER BY created_at DESC
    """)

    for row in cursor.fetchall():
        orders_sheet.append(tuple(row))

    conn.close()

# ---------------- SAVE FILE ----------------
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="restaurant_reports.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/owner/export/pdf")
def export_pdf():
    conn = get_db_connection()
    cursor = conn.cursor()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # ------------------- TITLE -------------------
    elements.append(Paragraph("Restaurant Report", styles["Title"]))
    elements.append(Spacer(1, 20))

    # ------------------- DAILY REVENUE -------------------
    cursor.execute("""
        SELECT DATE(created_at),
               COUNT(id),
               SUM(total)
        FROM orders
        WHERE status='completed'
        GROUP BY DATE(created_at)
        ORDER BY DATE(created_at) DESC
    """)
    rows = cursor.fetchall()
    daily_data = [["Date", "Orders", "Revenue"]]
    for r in rows:
        daily_data.append([r[0], r[1], f"Rs.{r[2] or 0}"])

    elements.append(Paragraph("Daily Revenue", styles["Heading2"]))
    daily_table = Table(daily_data, hAlign="LEFT", colWidths=[100, 60, 80])
    daily_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN',(1,1),(-1,-1),'CENTER'),
    ]))
    elements.append(daily_table)
    elements.append(Spacer(1,20))
    
    # ------------------- MONTHLY REVENUE -------------------
    # cursor.execute("""
    #     SELECT strftime('%Y-%m', created_at),
    #            COUNT(id),
    #            SUM(total)
    #     FROM orders
    #     WHERE status='completed'
    #     GROUP BY strftime('%Y-%m', created_at)
    #     ORDER BY strftime('%Y-%m', created_at) DESC
    # """)
    # rows = cursor.fetchall()
    # monthly_data = [["Month", "Orders", "Revenue"]]
    # for r in rows:
    #     monthly_data.append([r[0], r[1], f"Rs.{r[2] or 0}"])

    # elements.append(Paragraph("Monthly Revenue", styles["Heading2"]))
    # monthly_table = Table(monthly_data, hAlign="LEFT", colWidths=[100, 60, 80])
    # monthly_table.setStyle(TableStyle([
    #     ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
    #     ('TEXTCOLOR', (0,0), (-1,0), colors.white),
    #     ('GRID', (0,0), (-1,-1), 1, colors.black),
    #     ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    #     ('ALIGN',(1,1),(-1,-1),'CENTER'),
    # ]))
    # elements.append(monthly_table)
    # elements.append(Spacer(1,20))

    # ------------------- CUSTOMER DETAILS -------------------
    cursor.execute("""
        SELECT id, name, email, phone
        FROM customers
    """)
    rows = cursor.fetchall()
    customer_data = [["ID", "Name", "Email", "Phone"]]
    for r in rows:
        customer_data.append([r[0], r[1], r[2], r[3]])

    elements.append(Paragraph("Customer Details", styles["Heading2"]))
    customer_table = Table(customer_data, hAlign="LEFT", colWidths=[60, 120, 150, 100])
    customer_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN',(1,1),(-1,-1),'CENTER'),
    ]))
    elements.append(customer_table)
    elements.append(Spacer(1,20))

    cursor.execute("""
            SELECT strftime('%Y-%m', created_at),
                COUNT(id),
                SUM(total)
            FROM orders
            WHERE status='completed'
            GROUP BY strftime('%Y-%m', created_at)
            ORDER BY strftime('%Y-%m', created_at) DESC
        """)
    rows = cursor.fetchall()
    monthly_data = [["Month", "Orders", "Revenue"]]
    for r in rows:
        monthly_data.append([r[0], r[1], f"Rs.{r[2] or 0}"])

    elements.append(Paragraph("Monthly Revenue", styles["Heading2"]))
    monthly_table = Table(monthly_data, hAlign="LEFT", colWidths=[100, 60, 80])
    monthly_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN',(1,1),(-1,-1),'CENTER'),
    ]))
    elements.append(monthly_table)
    elements.append(Spacer(1,20))
    # ------------------- BUILD PDF -------------------
    doc.build(elements)
    buffer.seek(0)
    conn.close()

    return send_file(
        buffer,
        as_attachment=True,
        download_name="restaurant_report.pdf",
        mimetype="application/pdf"
    )
    # -------------------
    # ORDER HISTORY
    # -------------------

    # cursor.execute("""
    #     SELECT id,
    #            table_no,
    #            total,
    #             paid,
    #            payment_method,
    #            created_at
    #     FROM orders
    #     ORDER BY created_at DESC
    # """)

    # rows = cursor.fetchall()

    # order_data = [["Order ID","Table","Amount","Paid","Payment_Method","Date"]]

    # for r in rows:
    #     order_data.append(list(r))

    # elements.append(Paragraph("Order History", styles["Heading2"]))
    # elements.append(Table(order_data))
    # elements.append(Spacer(1,20))

    
    # # -------------------
    # # BUILD PDF
    # # -------------------

    # doc.build(elements)

    # buffer.seek(0)

    # conn.close()

    # return send_file(
    #     buffer,
    #     as_attachment=True,
    #     download_name="restaurant_report.pdf",
    #     mimetype="application/pdf"
    # )

@app.route("/owner/report/summary")
def report_summary():

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT COUNT(id), SUM(total)
        FROM orders
        WHERE status='completed'
    """)

    result = cursor.fetchone()

    conn.close()

    return jsonify({
        "total_orders": result[0] or 0,
        "total_revenue": result[1] or 0
    })

@app.route("/owner/report/daily")
def report_daily():

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT DATE(created_at),
               COUNT(id),
               SUM(total)
        FROM orders
        WHERE status='completed'
        GROUP BY DATE(created_at)
        ORDER BY DATE(created_at) DESC
    """)

    rows = cursor.fetchall()

    data = []

    for r in rows:
        data.append({
            "date": r[0],
            "orders": r[1],
            "revenue": r[2]
        })

    conn.close()

    return jsonify(data)

@app.route("/owner/report/monthly")
def report_monthly():

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT strftime('%Y-%m', created_at),
               COUNT(id),
               SUM(total)
        FROM orders
        WHERE status='completed'
        GROUP BY strftime('%Y-%m', created_at)
        ORDER BY strftime('%Y-%m', created_at) DESC
    """)

    rows = cursor.fetchall()

    data = []

    for r in rows:
        data.append({
            "month": r[0],
            "orders": r[1],
            "revenue": r[2]
        })

    conn.close()

    return jsonify(data)

@app.route("/owner/report/customers")
def report_customers():

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT id,name,email,phone FROM customers")

    rows = cursor.fetchall()

    data = []

    for r in rows:
        data.append({
            "id": r[0],
            "name": r[1],
            "email": r[2],
            "phone": r[3]
        })

    conn.close()

    return jsonify(data)

@app.route("/owner/report/orders")
def report_orders():
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT id, table_no, total, paid, payment_method, created_at
            FROM orders
            ORDER BY created_at DESC
        """)

        rows = cursor.fetchall()

        data = []
        for r in rows:
            data.append({
                "id": r[0],
                "table_no": r[1],
                "total": r[2],
                "paid": r[3],
                "payment_method": r[4] if len(r) > 4 else None,
                "created_at": r[5]
            })

        print("ORDERS DATA:", data)  # 👈 DEBUG

        return jsonify(data)

    except Exception as e:
        print("ERROR:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        conn.close()
#----------------------------------------
# Owner Insights
#----------------------------------------

# 10) Top customer
@app.route("/owner/top-customer", methods=["GET"])
def owner_top_customer():
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT c.id, c.name,
                   SUM(o.total) AS total_spent
            FROM customers c
            JOIN orders o ON c.id = o.user_id
            WHERE o.paid = 1
            GROUP BY c.id
            ORDER BY total_spent DESC
            LIMIT 1
        """)
        
        row = cur.fetchone()
        conn.close()

        if not row:
            return jsonify({"message": "No customers found", "data": None}), 200

        return jsonify({
            "id": row["id"],
            "name": row["name"],
            "total_spent": row["total_spent"]
        })

    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500

# 11) Top category (by order_items JOIN menu_items)
@app.route("/owner/top-category", methods=["GET"])
def owner_top_category():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT mi.category AS category, COALESCE(SUM(oi.quantity * oi.price),0) AS revenue
            FROM order_items oi
            LEFT JOIN menu_items mi ON oi.menu_id = mi.id
            GROUP BY mi.category
            ORDER BY revenue DESC
            LIMIT 1
        """)
        row = cur.fetchone()
        conn.close()
        if row:
            return jsonify({"category": row["category"] or "N/A", "revenue": float(row["revenue"] or 0)}), 200
        return jsonify({"category": "N/A", "revenue": 0}), 200
    except Exception as e:
        print("Top category error:", e)
        return jsonify({"error": "Server error"}), 500

# 12) Top table
@app.route("/owner/top-table", methods=["GET"])
def owner_top_table():
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT table_no,
                   COUNT(id) AS total_orders
            FROM orders
            GROUP BY table_no
            ORDER BY total_orders DESC
            LIMIT 1
        """)
        
        row = cur.fetchone()
        conn.close()

        if not row:
            return jsonify({"message": "No table data", "data": None}), 200

        return jsonify({
            "table_no": row["table_no"],
            "total_orders": row["total_orders"]
        })

    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 500

# 13) Peak hour
@app.route("/owner/peak-hour", methods=["GET"])
def owner_peak_hour():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT strftime('%H', created_at) as hour,
                   COUNT(*) as cnt
            FROM orders
            GROUP BY hour
            ORDER BY cnt DESC
            LIMIT 1
        """)
        row = cur.fetchone()
        conn.close()
        if row:
            return jsonify({"hour": row["hour"], "orders": int(row["cnt"])}), 200
        return jsonify({"hour": "N/A", "orders": 0}), 200
    except Exception as e:
        print("Peak hour error:", e)
        return jsonify({"error": "Server error"}), 500

# 14) Profit (simple)
@app.route("/owner/profit", methods=["GET"])
def owner_profit():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT COALESCE(SUM(COALESCE(totalAmount, total)),0) as rev FROM orders WHERE status = 'completed'")
        r = cur.fetchone()
        conn.close()
        revenue = float(r["rev"] or 0)
        profit = revenue * 0.30
        return jsonify({"revenue": revenue, "profit": profit}), 200
    except Exception as e:
        print("Profit error:", e)
        return jsonify({"error": "Server error"}), 500

# 15) Export all orders CSV
@app.route("/owner/export", methods=["GET"])
def owner_export_orders():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        # include both totalAmount and total for compatibility
        cur.execute("SELECT id, user_id, COALESCE(totalAmount, total) as total, status, created_at FROM orders")
        rows = cur.fetchall()
        conn.close()

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["OrderID", "UserID", "Total", "Status", "Created At"])
        for r in rows:
            writer.writerow([r["id"], r["user_id"], r["total"], r["status"], r["created_at"]])

        return Response(output.getvalue(), mimetype="text/csv",
                        headers={"Content-disposition": "attachment; filename=orders.csv"})
    except Exception as e:
        print("owner_export_orders error:", e)
        return jsonify({"error": "Server error"}), 500

# 16) OPTIONAL: Manager / public endpoint returns normalized orders for frontend
# If you don't have this already, add /manager/orders that returns orders in the exact shape frontend expects:
@app.route("/owner/order-items/<int:order_id>", methods=["GET"])
def owner_get_order_items(order_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT 
                item_name AS name,
                quantity,
                price
            FROM order_items
            WHERE order_id = %s
        """, (order_id,))

        items = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify(items), 200

    except Exception as e:
        print("order-items error:", e)
        return jsonify([]), 200

from flask import send_file
import io, csv, json
from datetime import datetime, timedelta

# --- Helper: convert sqlite rows to dict list (if not present) ---
def rows_to_list(rows):
    return [dict(r) for r in rows] if rows else []

# 1) Sales: flexible range (daily, weekly, monthly)
@app.route("/owner/sales", methods=["GET"])
def owner_sales():
    # query params: %srange=daily|weekly|monthly|custom&start=YYYY-MM-DD&end=YYYY-MM-DD
    rng = request.args.get("range","daily")
    start = request.args.get("start")
    end = request.args.get("end")
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        if rng == "daily":
            cur.execute("""
              SELECT DATE(created_at) as day, COALESCE(SUM(total),0) as total
              FROM orders
              GROUP BY DATE(created_at)
              ORDER BY DATE(created_at) DESC
              LIMIT 30
            """)
            rows = cur.fetchall()
            data = [{"day": r["day"], "total": float(r["total"])} for r in rows]
        elif rng == "weekly":
            cur.execute("""
              SELECT strftime('%Y-%W', created_at) as week, COALESCE(SUM(total),0) as total
              FROM orders
              GROUP BY week ORDER BY week DESC LIMIT 26
            """)
            rows = cur.fetchall()
            data = [{"week": r["week"], "total": float(r["total"])} for r in rows]
        elif rng == "monthly":
            cur.execute("""
              SELECT strftime('%Y-%m', created_at) as month, COALESCE(SUM(total),0) as total
              FROM orders
              GROUP BY month ORDER BY month DESC LIMIT 24
            """)
            rows = cur.fetchall()
            data = [{"month": r["month"], "total": float(r["total"])} for r in rows]
        elif rng == "custom" and start and end:
            cur.execute("""
              SELECT DATE(created_at) as day, COALESCE(SUM(total),0) as total
              FROM orders
              WHERE DATE(created_at) BETWEEN %s AND %s
              GROUP BY DATE(created_at) ORDER BY DATE(created_at) ASC
            """, (start, end))
            rows = cur.fetchall()
            data = [{"day": r["day"], "total": float(r["total"])} for r in rows]
        else:
            data = []
        conn.close()
        return jsonify(data), 200
    except Exception as e:
        print("owner_sales error:", e)
        return jsonify([]), 200

# 3) Staff performance (basic: completed orders count)
@app.route("/owner/staff/performance", methods=["GET"])
def owner_staff_performance():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT user_id, COUNT(*) as completed_orders, SUM(total) as revenue
            FROM orders
            WHERE status='completed'
            GROUP BY user_id
            ORDER BY completed_orders DESC
            LIMIT 20
        """)
        rows = cur.fetchall()
        conn.close()
        # Map user_id -> name if you keep customers table or staff table, but return data raw
        return jsonify([dict(r) for r in rows]), 200
    except Exception as e:
        print("staff performance error:", e)
        return jsonify([]), 200

# 4) Table status (occupied/free counts)
@app.route("/owner/table-status", methods=["GET"])
def owner_table_status():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        # active orders (not completed/cancelled) per table
        cur.execute("""
            SELECT table_no, COUNT(*) as active_orders
            FROM orders
            WHERE status IN ('pending','preparing','ready')
            GROUP BY table_no
        """)
        rows = cur.fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows]), 200
    except Exception as e:
        print("table status error:", e)
        return jsonify([]), 200

# 5) Notifications: historical notifications table (optional), simple endpoint
@app.route("/owner/notifications", methods=["GET"])
def owner_notifications():
    try:
        # For now, create notifications on events via socket and store in memory or db.
        conn = get_db_connection()
        cur = conn.cursor()
        # if you want a persistent table: notifications(id, message, type, created_at)
        # Here return last N from a notifications table if exists, else return empty
        try:
            cur.execute("SELECT id, message, type, created_at FROM notifications ORDER BY created_at DESC LIMIT 50")
            rows = cur.fetchall()
            conn.close()
            return jsonify([dict(r) for r in rows]), 200
        except Exception:
            conn.close()
            return jsonify([]), 200
    except Exception as e:
        print("notifications error:", e)
        return jsonify([]),200

# 6) All orders (for 'view all'): return enriched order rows with items
@app.route("/owner/orders/all", methods=["GET"])
def owner_all_orders():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, user_id, table_no, total, paid, status, created_at FROM orders ORDER BY created_at DESC")
        orders = [dict(r) for r in cur.fetchall()]
        for o in orders:
            cur.execute("SELECT oi.menu_id, oi.quantity, oi.price, mi.name as item_name FROM order_items oi LEFT JOIN menu_items mi ON mi.id = oi.menu_id WHERE oi.order_id = %s", (o["id"],))
            items = [dict(x) for x in cur.fetchall()]
            o["items"] = items
        conn.close()
        return jsonify(orders), 200
    except Exception as e:
        print("owner_all_orders error:", e)
        return jsonify([]),200

# -------------------------------------------------
# OWNER DASHBOARD REVENUE ANALYTICS
# -------------------------------------------------
@app.route("/owner/analytics/revenue", methods=["GET"])
def owner_revenue_analytics():
    range_type = request.args.get("range", "week")  # day / week / month
    conn = get_db_connection()
    cur = conn.cursor()
    data = []

    try:
        if range_type == "day":
            cur.execute("""
                SELECT strftime('%H', created_at) AS hour,
                       SUM(total / (1 + 0.05 + 0.05)) AS net_revenue,
                       SUM(total) AS gross_revenue
                FROM orders
                WHERE paid = 1 AND status = 'completed'
                      AND DATE(created_at) = DATE('now')
                GROUP BY hour
                ORDER BY hour
            """)
            rows = cur.fetchall()
            data = [{"label": f"{r['hour']}:00", "net": r["net_revenue"] or 0, "gross": r["gross_revenue"] or 0} for r in rows]

        elif range_type == "week":
            cur.execute("""
                SELECT strftime('%w', created_at) AS weekday,
                       SUM(total / (1 + 0.05 + 0.05)) AS net_revenue,
                       SUM(total) AS gross_revenue
                FROM orders
                WHERE paid = 1 AND status = 'completed'
                      AND DATE(created_at) >= DATE('now', '-6 days')
                GROUP BY weekday
                ORDER BY weekday
            """)
            rows = cur.fetchall()
            weekmap = {"0":"Sun","1":"Mon","2":"Tue","3":"Wed","4":"Thu","5":"Fri","6":"Sat"}
            data = [{"label": weekmap[r["weekday"]], "net": r["net_revenue"] or 0, "gross": r["gross_revenue"] or 0} for r in rows]

        elif range_type == "month":
            cur.execute("""
                SELECT strftime('%d', created_at) AS day,
                       SUM(total / (1 + 0.05 + 0.05)) AS net_revenue,
                       SUM(total) AS gross_revenue
                FROM orders
                WHERE paid = 1 AND status = 'completed'
                      AND strftime('%Y-%m', created_at) = strftime('%Y-%m', 'now')
                GROUP BY day
                ORDER BY day
            """)
            rows = cur.fetchall()
            data = [{"label": r["day"], "net": r["net_revenue"] or 0, "gross": r["gross_revenue"] or 0} for r in rows]

        conn.close()
        return jsonify(data)

    except Exception as e:
        conn.close()
        print("owner revenue analytics error:", e)
        return jsonify([])
    
# -------------------------------
# CATEGORY REVENUE DISTRIBUTION
# -------------------------------
# @app.route("/owner/analytics/category-revenue")
# def category_revenue():
#     conn = get_db_connection()
#     cur = conn.cursor()

#     cur.execute("""
#         SELECT m.category, SUM(oi.quantity * oi.price)
#         FROM order_items oi
#         JOIN menu_items m ON m.id = oi.menu_id
#         JOIN orders o ON o.id = oi.order_id
#         WHERE o.status = 'completed'
#         GROUP BY m.category
#     """)

#     rows = cur.fetchall()
#     conn.close()

#     return jsonify([
#         {"name": name, "value": value or 0}
#         for name, value in rows
#     ])

@app.route("/owner/today-report")
def today_report():

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT 
                o.id,
                c.name AS customer_name,

                STRING_AGG(
                    m.name || ' (' || oi.quantity || ')', ', '
                ) AS items,

                o.total AS amount,
                o.paid,
                o.created_at

            FROM orders o

            LEFT JOIN customers c
            ON o.user_id = c.id

            JOIN order_items oi
            ON o.id = oi.order_id

            JOIN menu_items m
            ON oi.menu_id = m.id

            WHERE DATE(o.created_at) = CURRENT_DATE

            GROUP BY o.id, c.name, o.total, o.paid, o.created_at

            ORDER BY o.created_at DESC
        """)

        rows = cur.fetchall()

        orders = []

        for r in rows:
            orders.append({
                "id": r["id"],
                "customer": r["customer_name"] or "Walk-in",
                "items": r["items"],
                "amount": float(r["amount"] or 0),
                "paid": "Paid" if r["paid"] else "Unpaid",
                "time": r["created_at"]
            })

        return jsonify(orders)

    finally:
        cur.close()
        conn.close()
        
# -----------------------------------------------------------------------------------------
# Analytics
#-------------------------------------------------------------------------------------------
@app.route("/analytics/orders/today", methods=["GET"])
def order_analytics_today():
    conn = get_db_connection()
    cur = conn.cursor()

    # Convert TEXT timestamp to DATE
    today = datetime.now().strftime("%Y-%m-%d")

    # 1️⃣ Today total orders
    cur.execute("""
        SELECT COUNT(*) 
        FROM orders 
        WHERE DATE(created_at) = DATE(%s)
    """, (today,))
    total_orders = cur.fetchone()[0]

    # 2️⃣ Today revenue (sum totalAmount where paid = 1)
    cur.execute("""
        SELECT SUM(total)
        FROM orders
        WHERE paid = 1
        AND DATE(created_at) = DATE(%s)
    """, (today,))
    revenue = cur.fetchone()[0] or 0

    # 3️⃣ Pending orders
    cur.execute("SELECT COUNT(*) FROM orders WHERE status='pending'")
    pending_orders = cur.fetchone()[0]

    # 4️⃣ Cancelled orders
    cur.execute("SELECT COUNT(*) FROM orders WHERE status='cancelled'")
    cancelled_orders = cur.fetchone()[0]

    conn.close()

    return jsonify({
        "today_orders": total_orders,
        "today_revenue": revenue,
        "pending_orders": pending_orders,
        "cancelled_orders": cancelled_orders
    })

@app.route("/orders/filter", methods=["GET"])
def filter_orders():
    status = request.args.get("status", None)
    table = request.args.get("table", None)
    customer = request.args.get("customer", None)
    date = request.args.get("date", None)  # today, yesterday, last7
    order_id = request.args.get("order_id", None)

    conn = get_db_connection()
    cur = conn.cursor()

    base = "SELECT * FROM orders WHERE 1=1"
    params = []

    if status:
        base += " AND status = %s"
        params.append(status)

    if table:
        base += " AND table_no = %s"
        params.append(table)

    if customer:
        base += " AND user_id = %s"
        params.append(customer)

    # --- Date filter logic ---
    if date == "today":
        base += " AND DATE(created_at) = DATE('now','localtime')"

    elif date == "yesterday":
        base += " AND DATE(created_at) = DATE('now','-1 day','localtime')"

    elif date == "last7":
        base += " AND DATE(created_at) >= DATE('now','-7 day','localtime')"

    if order_id:
        base += " AND id = %s"
        params.append(order_id)

    cur.execute(base, params)
    rows = cur.fetchall()

    conn.close()

    return jsonify([dict(r) for r in rows])

@app.route("/analytics/staff", methods=["GET"])
def analytics_staff():
    conn = get_db_connection()
    cur = conn.cursor()

    # Count total staff per role
    cur.execute("SELECT COUNT(*) FROM managers")
    total_managers = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM chefs")
    total_chefs = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM waiters")
    total_waiters = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM owners")
    total_owners = cur.fetchone()[0]

    # Count active staff (only those tables which have 'active')
    cur.execute("SELECT COUNT(*) FROM managers WHERE active = 1")
    active_managers = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM chefs WHERE active = 1")
    active_chefs = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM waiters WHERE active = 1")
    active_waiters = cur.fetchone()[0]

    # owners don't have active column → assume all are active
    active_owners = total_owners

    active_total = active_managers + active_chefs + active_waiters + active_owners
    inactive_total = (total_managers + total_chefs + total_waiters) - (active_managers + active_chefs + active_waiters)

    conn.close()

    return jsonify({
        "managers": total_managers,
        "chefs": total_chefs,
        "waiters": total_waiters,
        "owners": total_owners,
        "active": active_total,
        "inactive": inactive_total
    })

@app.route("/analytics/tables", methods=["GET"])
def table_occupancy():
    conn = get_db_connection()
    cur = conn.cursor()

    # 1. Get all tables from existing orders (unique table_no)
    cur.execute("SELECT DISTINCT table_no FROM orders WHERE table_no IS NOT NULL")
    table_rows = cur.fetchall()

    tables = [row["table_no"] for row in table_rows]

    result = []

    for t in tables:
        # Check active order for this table (not paid, not cancelled)
        cur.execute("""
            SELECT id, status, paid 
            FROM orders 
            WHERE table_no = %s 
            ORDER BY id DESC LIMIT 1
        """, (t,))
        order = cur.fetchone()

        if not order:
            result.append({"table_no": t, "status": "free"})
            continue

        # Determine status
        if order["status"] in ("pending", "preparing", "ready"):
            state = "occupied"
        elif order["paid"] == 0 and order["status"] == "completed":
            state = "waiting_for_bill"
        elif order["status"] == "cancelled":
            state = "free"
        else:
            state = "free"

        result.append({
            "table_no": t,
            "status": state,
            "order_id": order["id"]
        })

    conn.close()

    return jsonify(result)

@app.route("/analytics/weekly", methods=["GET"])
def weekly_sales():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            strftime('%w', created_at) AS weekday,
            COUNT(*) AS total_orders,
            SUM(total) AS revenue
        FROM orders
        GROUP BY weekday
        ORDER BY weekday
    """)

    rows = cur.fetchall()
    conn.close()

    # Map numbers to weekday names
    weekmap = {
        "0": "Sunday",
        "1": "Monday",
        "2": "Tuesday",
        "3": "Wednesday",
        "4": "Thursday",
        "5": "Friday",
        "6": "Saturday"
    }

    result = []
    for r in rows:
        result.append({
            "day": weekmap[r["weekday"]],
            "orders": r["total_orders"],
            "revenue": r["revenue"] or 0
        })

    return jsonify(result)


@app.route("/analytics/staff_attendance", methods=["GET"])
def staff_attendance():
    conn = get_db_connection()
    cur = conn.cursor()

    # Count present chefs
    cur.execute("SELECT COUNT(*) AS count FROM chefs WHERE active = 1")
    chefs_present = cur.fetchone()["count"]

    cur.execute("SELECT COUNT(*) AS count FROM chefs WHERE active = 0")
    chefs_absent = cur.fetchone()["count"]

    # Count present waiters
    cur.execute("SELECT COUNT(*) AS count FROM waiters WHERE active = 1")
    waiters_present = cur.fetchone()["count"]

    cur.execute("SELECT COUNT(*) AS count FROM waiters WHERE active = 0")
    waiters_absent = cur.fetchone()["count"]

    # Count present managers
    cur.execute("SELECT COUNT(*) AS count FROM managers WHERE active = 1")
    managers_present = cur.fetchone()["count"]

    cur.execute("SELECT COUNT(*) AS count FROM managers WHERE active = 0")
    managers_absent = cur.fetchone()["count"]

    conn.close()

    return jsonify({
        "chefs": {"present": chefs_present, "absent": chefs_absent},
        "waiters": {"present": waiters_present, "absent": waiters_absent},
        "managers": {"present": managers_present, "absent": managers_absent}
    })

# @app.route("/inventory", methods=["GET"])
# def get_inventory():
#     conn = get_db_connection()
#     items = conn.execute("SELECT * FROM inventory ORDER BY item_name").fetchall()
#     conn.close()
#     return jsonify([dict(row) for row in items])

# @app.route("/inventory", methods=["POST"])
# def add_inventory_item():
#     data = request.json
#     name = data.get("item_name")
#     sku = data.get("sku")
#     qty = data.get("quantity")
#     unit = data.get("unit")
#     threshold = data.get("low_threshold")
#     note = data.get("note")

#     conn = get_db_connection()
#     conn.execute("""
#         INSERT INTO inventory (item_name, sku, quantity, unit, low_threshold, note)
#         VALUES (%s, %s, %s, %s, %s, %s)
#     """, (name, sku, qty, unit, threshold, note))

#     conn.commit()
#     conn.close()
#     return jsonify({"message": "Inventory item added"})

# @app.route("/inventory/<int:item_id>", methods=["PUT"])
# def update_inventory(item_id):
#     data = request.json
#     qty = data.get("quantity")
#     note = data.get("note")

#     conn = get_db_connection()
#     conn.execute("""
#         UPDATE inventory 
#         SET quantity = %s, updated_at = datetime('now'), note = %s
#         WHERE id = %s
#     """, (qty, note, item_id))

#     conn.commit()
#     conn.close()
#     return jsonify({"message": "Inventory updated"})

# @app.route("/inventory/<int:item_id>", methods=["DELETE"])
# def delete_inventory(item_id):
#     conn = get_db_connection()
#     conn.execute("DELETE FROM inventory WHERE id = %s", (item_id,))
#     conn.commit()
#     conn.close()
#     return jsonify({"message": "Item deleted"})

# @app.route("/inventory/low", methods=["GET"])
# def low_stock_inventory():
#     conn = get_db_connection()
#     items = conn.execute("""
#         SELECT * FROM inventory
#         WHERE quantity <= low_threshold
#     """).fetchall()

#     conn.close()
#     return jsonify([dict(row) for row in items])

# -----------------------------
# EXPENSE MANAGER API
# -----------------------------

@app.route("/expenses", methods=["GET"])
def get_expenses():
    conn = get_db_connection()
    expenses = conn.execute("SELECT * FROM expenses ORDER BY created_at DESC").fetchall()
    conn.close()

    data = [dict(row) for row in expenses]
    return jsonify({"success": True, "expenses": data})


@app.route("/expenses/add", methods=["POST"])
def add_expense():
    data = request.json
    title = data.get("title")
    category = data.get("category")
    amount = data.get("amount")
    note = data.get("note", "")

    conn = get_db_connection()
    conn.execute(
        "INSERT INTO expenses (title, category, amount, note) VALUES (%s, %s, %s, %s)",
        (title, category, amount, note)
    )
    conn.commit()
    conn.close()

    return jsonify({"success": True, "message": "Expense added"})


@app.route("/expenses/delete/<int:expense_id>", methods=["DELETE"])
def delete_expense(expense_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM expenses WHERE id=%s", (expense_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Expense deleted"})


@app.route("/expenses/summary", methods=["GET"])
def expense_summary():
    conn = get_db_connection()

    today = conn.execute("""
        SELECT SUM(amount) as total FROM expenses
        WHERE DATE(created_at) = DATE('now')
    """).fetchone()

    monthly = conn.execute("""
        SELECT SUM(amount) as total FROM expenses
        WHERE strftime('%Y-%m', created_at) = strftime('%Y-%m','now')
    """).fetchone()

    conn.close()

    return jsonify({
        "success": True,
        "today_expense": today["total"] or 0,
        "monthly_expense": monthly["total"] or 0
    })

@app.route("/analytics/customer/frequent")
def most_frequent_customer():
    conn = get_db_connection()
    row = conn.execute("""
        SELECT customer_name, COUNT(*) as visits
        FROM orders
        WHERE customer_name != ''
        GROUP BY customer_name
        ORDER BY visits DESC
        LIMIT 1
    """).fetchone()
    conn.close()

    return jsonify({"success": True, "data": dict(row) if row else None})

@app.route("/analytics/customer/highest")
def highest_paying_customer():
    conn = get_db_connection()
    row = conn.execute("""
        SELECT customer_name, SUM(COALESCE(totalAmount, total)) AS spent
        FROM orders
        WHERE customer_name != ''
        GROUP BY customer_name
        ORDER BY spent DESC
        LIMIT 1
    """).fetchone()
    conn.close()

    return jsonify({"success": True, "data": dict(row) if row else None})

@app.route("/analytics/customer/today")
def customers_today():
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT DISTINCT customer_name
        FROM orders
        WHERE DATE(created_at) = DATE('now')
          AND customer_name != ''
    """).fetchall()
    conn.close()

    return jsonify({"success": True, "data": [r["customer_name"] for r in rows]})

@app.route("/analytics/customer/unpaid")
def customers_unpaid():
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT DISTINCT customer_name
        FROM orders
        WHERE paid = 0 AND customer_name != ''
    """).fetchall()
    conn.close()

    return jsonify({"success": True, "data": [r["customer_name"] for r in rows]})

@app.route("/analytics/customer/full")
def customer_full_analytics():
    conn = get_db_connection()
    c = conn.cursor()

    # --- 1. Purchase history per day (Daily Revenue) ---
    c.execute("""
        SELECT 
            DATE(created_at) AS date,
            SUM(COALESCE(totalAmount, total)) AS amount
        FROM orders
        WHERE paid = 1
        GROUP BY DATE(created_at)
        ORDER BY date ASC
    """)
    daily_rows = c.fetchall()
    daily_history = [{"date": r["date"], "amount": r["amount"]} for r in daily_rows]

    # --- 2. Customer Purchase Totals ---
    c.execute("""
        SELECT 
            customer_name,
            SUM(COALESCE(totalAmount, total)) AS total_spent
        FROM orders
        WHERE paid = 1
        GROUP BY customer_name
        HAVING customer_name IS NOT NULL AND customer_name != ''
        ORDER BY total_spent DESC
    """)
    customer_rows = c.fetchall()
    purchase_history = [
        {
            "customer_name": r["customer_name"],
            "total_spent": r["total_spent"]
        }
        for r in customer_rows
    ]

    # --- 3. Most Frequent Customer ---
    c.execute("""
        SELECT customer_name, COUNT(*) as visits
        FROM orders
        WHERE customer_name IS NOT NULL AND customer_name != ''
        GROUP BY customer_name
        ORDER BY visits DESC
        LIMIT 1
    """)
    most_frequent = c.fetchone()
    frequent_customer = {
        "customer_name": most_frequent["customer_name"] if most_frequent else None,
        "visits": most_frequent["visits"] if most_frequent else 0
    }

    # --- 4. Highest Paying Customer ---
    highest_paying = purchase_history[0] if len(purchase_history) > 0 else None

    # --- 5. Customers visited today ---
    c.execute("""
        SELECT DISTINCT customer_name
        FROM orders
        WHERE DATE(created_at) = DATE('now')
        AND customer_name IS NOT NULL AND customer_name != ''
    """)
    today_visits = [row["customer_name"] for row in c.fetchall()]

    # --- 6. Customers with Unpaid Bills ---
    c.execute("""
        SELECT customer_name, SUM(total) AS due
        FROM orders
        WHERE paid = 0
        AND customer_name IS NOT NULL AND customer_name != ''
        GROUP BY customer_name
    """)
    unpaid_rows = c.fetchall()
    unpaid_customers = [
        {"customer_name": r["customer_name"], "due": r["due"]}
        for r in unpaid_rows
    ]

    conn.close()

    # --- Final Combined JSON ---
    return jsonify({
        "daily_history": daily_history,
        "purchase_history": purchase_history,
        "most_frequent_customer": frequent_customer,
        "highest_paying_customer": highest_paying,
        "customers_today": today_visits,
        "unpaid_customers": unpaid_customers
    })

# -----------------------------------
# OWNER: CUSTOMER PURCHASE HISTORY
# -----------------------------------
# -----------------------------
# Helper: Fetch Purchase History
# -----------------------------
def get_purchase_history():
    conn = get_db_connection()
    c = conn.cursor()

    try:
        c.execute("""
            SELECT 
                COALESCE(c.name, 'Guest Customer') AS customer_name,
                SUM(o.total) AS total_spent,
                COUNT(o.id) AS visits
            FROM orders o
            LEFT JOIN customers c ON o.user_id = c.id
            WHERE o.paid = 1
            GROUP BY c.id, c.name
            ORDER BY total_spent DESC
        """)

        rows = c.fetchall()

        history = [
            {
                "customer_name": row["customer_name"],
                "total_spent": float(row["total_spent"] or 0),
                "visits": int(row["visits"] or 0)
            }
            for row in rows
        ]

        return history

    finally:
        c.close()
        conn.close()


@app.route("/owner/purchase-history")
def owner_purchase_history():
    try:
        history = get_purchase_history()
        return jsonify({"history": history}), 200
    except Exception as e:
        print("ERROR in /owner/purchase-history:", e)
        return jsonify({"error": "server error"}), 500


@app.route("/analytics/customer/history")
def analytics_purchase_history():
    try:
        history = get_purchase_history()
        return jsonify({"history": history}), 200
    except Exception as e:
        print("ERROR in /analytics/customer/history:", e)
        return jsonify({"error": "server error"}), 500


# ---------------------------
# Start server
# ---------------------------
if __name__ == "__main__":
    import os

    # ✅ Initialize DB (PostgreSQL)
    init_db()

    # ✅ Optional: run seeds only once (recommended)
    if os.getenv("RUN_SEED", "false") == "true":
        seed_menu_items()
        seed_tables()

    # ✅ Get dynamic port from Render
    port = int(os.environ.get("PORT", 5000))

    # ✅ Run server (IMPORTANT CHANGE)
    socketio.run(app, host="0.0.0.0", port=port)